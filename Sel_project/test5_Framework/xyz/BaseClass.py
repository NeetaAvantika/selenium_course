import inspect
import logging

import pytest
from selenium.webdriver.common.by import By
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.select import Select
import os

import openpyxl

@pytest.mark.usefixtures("setup")
class BaseClass:

    def getTestData(self,test_case_name):
        Dict = {}
        book = openpyxl.load_workbook(os.getcwd() + "\\TestData\\Gmail_send_mail_Testdata.xlsx")
        sheet = book.active
        for i in range(1, sheet.max_row + 1):  # to get rows
            if sheet.cell(row=i, column=1).value == test_case_name:

                for j in range(2, sheet.max_column + 1):  # to get columns
                    # Dict["lastname"]="shetty
                    Dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value
        return Dict

    def getLogger(self):
        loggerName = inspect.stack()[1][3]
        logger = logging.getLogger(loggerName)
        fileHandler = logging.FileHandler('logfile.log')
        formatter = logging.Formatter("%(asctime)s :%(levelname)s : %(name)s :%(message)s")
        fileHandler.setFormatter(formatter)

        logger.addHandler(fileHandler)  # filehandler object

        logger.setLevel(logging.DEBUG)
        return logger

    def verifyLinkPresence(self, text):
        element = WebDriverWait(self.driver, 10).until(
        EC.presence_of_element_located((By.LINK_TEXT, text)))

    def selectOptionByText(self,locator,text):
        sel = Select(locator)
        sel.select_by_visible_text(text)
